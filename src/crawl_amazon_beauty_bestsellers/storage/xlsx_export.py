from __future__ import annotations

import json
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..config import Settings
from .store import Store


def _autosize(ws):
    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 8), 60)


CURATED_SPEC_COLUMNS = [
    ("Item Form", "item_form"),
    ("Skin Type", "skin_type"),
    ("Item Weight", "item_weight"),
    ("Unit Count", "unit_count"),
    ("Active Ingredients", "active_ingredients"),
    ("Material Type Free", "material_free"),
    ("Sun Protection Factor", "spf"),
    ("Age Range Description", "age_range"),
    ("Country as Labeled", "country"),
    ("Country of Origin", "country"),
    ("Recommended Uses For Product", "recommended_uses"),
    ("Product Benefits", "product_benefits"),
]


def _parse_specs(raw: str) -> dict[str, str]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except (json.JSONDecodeError, TypeError):
        return {}


def export_day(settings: Settings, store: Store, date_str: str | None = None, name_map: dict[str, str] | None = None) -> Path:
    date_str = date_str or time.strftime("%Y-%m-%d")
    wb = Workbook()
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(bold=True, color="0563C1", underline="single")
    name_map = name_map or {}

    def _sheet_title(node_id: str) -> str:
        base = name_map.get(str(node_id)) or f"node_{node_id}"
        for ch in '[]:*?/\\':
            base = base.replace(ch, " ")
        return base[:31]

    index_ws = wb.active
    index_ws.title = "INDEX"
    index_ws.append(["node_id", "category", "bestsellers_url", "listed_rows"])
    for cell in index_ws[1]:
        cell.font = header_font

    def _q(title: str) -> str:
        return "'" + title.replace("'", "''") + "'"

    panel_titles: list[tuple[str, str]] = []
    per_node = store.day_latest_rows(date_str)
    for node_id, rows in sorted(per_node.items()):
        sheet_name = _sheet_title(node_id)
        panel_titles.append((sheet_name, str(node_id)))
        ws = wb.create_sheet(sheet_name)
        bestsellers_url = f"https://www.amazon.com/Best-Sellers/zgbs/beauty/{node_id}"
        ws.append([f"Amazon Best Sellers — {sheet_name}", "", "", "", "", "", "", "", bestsellers_url])
        link_cell = ws.cell(row=1, column=9)
        link_cell.value = "🔗 Best Sellers 페이지 열기"
        link_cell.hyperlink = bestsellers_url
        link_cell.font = link_font
        ws.cell(row=1, column=1).font = header_font
        headers = ["rank", "asin", "title", "rating", "ratings_count", "price", "currency", "offers_text", "url"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
        for row in rows:
            ws.append([
                row.get("rank"), row.get("asin"), row.get("title"),
                row.get("rating"), row.get("ratings_count"),
                row.get("price_amount"), row.get("price_currency"),
                row.get("offers_text"), row.get("url"),
            ])
        _autosize(ws)

    for title, node_id in panel_titles:
        index_ws.append([
            node_id, title,
            f"https://www.amazon.com/Best-Sellers/zgbs/beauty/{node_id}",
            f"=COUNT({_q(title)}!A:A)",
        ])
    _autosize(index_ws)

    def _category_formula(asin_ref: str, row_num: int) -> str:
        parts = ",".join(
            f'IF(COUNTIF({_q(t)}!$B$1:$B$500,{asin_ref}),"{t}","")'
            for t, _ in panel_titles
        )
        return f'=TEXTJOIN(" | ",TRUE,{parts})'

    details = store.detail_day_rows(date_str)
    if details:
        asin_nodes = store.day_asin_categories(date_str)
        parsed_details = [(row, _parse_specs(row.get("specs") or "")) for row in details]

        ws = wb.create_sheet("details")
        keys = [
            "brand", "manufacturer", "model_number", "seller_name",
            "buy_box_price", "buy_box_currency", "list_price_amount",
            "bsr_main_rank", "bsr_main_category",
            "bsr_sub_1_rank", "bsr_sub_1_category", "bsr_sub_2_rank", "bsr_sub_2_category",
            "ratings_histogram", "date_first_available",
            "availability", "price_source", "variants_count",
        ]
        header_cols: list[str] = []
        for _, column in CURATED_SPEC_COLUMNS:
            if column not in header_cols:
                header_cols.append(column)
        ws.append(["fetched_at", "asin", "category(자동)", "ranked_node_ids"] + keys + header_cols + ["specs_json"])
        for cell in ws[1]:
            cell.font = header_font

        def _flatten(row: dict) -> dict:
            flat = dict(row)
            try:
                bsr_other = json.loads(row.get("bsr_other") or "[]")
            except (json.JSONDecodeError, TypeError):
                bsr_other = []
            for idx in range(2):
                item = bsr_other[idx] if idx < len(bsr_other) else {}
                flat[f"bsr_sub_{idx + 1}_rank"] = item.get("rank")
                flat[f"bsr_sub_{idx + 1}_category"] = item.get("category")
            try:
                histogram = json.loads(row.get("ratings_histogram") or "{}")
            except (json.JSONDecodeError, TypeError):
                histogram = {}
            flat["ratings_histogram"] = " ".join(
                f"{star}:{pct}%" for star, pct in sorted(histogram.items(), reverse=True)
            )
            return flat

        for row, specs in parsed_details:
            row = _flatten(row)
            curated_values = []
            for column in header_cols:
                value = ""
                for label, target in CURATED_SPEC_COLUMNS:
                    if target != column:
                        continue
                    candidate = specs.get(label)
                    if candidate:
                        value = candidate
                        break
                curated_values.append(value)
            specs_pretty = json.dumps(specs, ensure_ascii=False)[:3000]
            r = ws.max_row + 1
            category_formula = _category_formula(f"$B{r}", r)
            node_ids = ", ".join(asin_nodes.get(str(row.get("asin")), []))
            ws.append(
                [row.get("fetched_at"), row.get("asin"), category_formula, node_ids]
                + [row.get(k) for k in keys] + curated_values + [specs_pretty]
            )
            ws.cell(row=ws.max_row, column=len(keys) + len(curated_values) + 2).alignment = wrap
        _autosize(ws)

        ws = wb.create_sheet("specs_long")
        ws.append(["asin", "category(자동)", "brand", "spec_key", "spec_value"])
        for cell in ws[1]:
            cell.font = header_font
        for row, specs in parsed_details:
            for key, value in sorted(specs.items()):
                r = ws.max_row + 1
                ws.append([
                    row.get("asin"), _category_formula(f"$A{r}", r),
                    row.get("brand"), key, str(value)[:500],
                ])
        _autosize(ws)

    trend = store.trend_rows(days=14)
    if trend:
        ws = wb.create_sheet("trend_14d")
        ws.append(["day", "node_id", "category(자동)", "asin", "best_rank", "snapshots", "max_ratings"])
        for cell in ws[1]:
            cell.font = header_font
        for row in trend:
            r = ws.max_row + 1
            values = list(row.values())
            category_formula = f'=IFERROR(VLOOKUP($B{r},INDEX!$A:$B,2,FALSE),"")'
            ws.append(values[:1] + [values[1], category_formula] + values[2:])
        _autosize(ws)

    exports_dir = settings.resolve(settings.storage.exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)
    out_path = exports_dir / f"amazon_bs_{date_str.replace('-', '')}.xlsx"
    wb.save(out_path)
    return out_path
